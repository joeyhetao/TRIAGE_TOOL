# -*- coding: utf-8 -*-
"""
KB 稳定 ID 迁移脚本（一次性，幂等）

对存量 error_db.xlsx：
  1. 缺 '稳定ID' 列时新加该列（写入表头 + 列宽）
  2. 任意行 '稳定ID' 单元格为空时，按 _compute_stable_id() 填入

第二次运行无变化（幂等）。返回 True 表示文件被改过。

不抛异常给调用方：失败时由 load_db() 的兜底吞掉，避免阻塞读路径。
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .db_manager import (
    _thread_lock, _FileLock, _save_atomic, ensure_db, _compute_stable_id,
)


def run(db_path: str) -> bool:
    """对 db_path 处的 KB 执行稳定 ID 迁移，返回是否真的改了文件。"""
    p = Path(db_path)
    if not p.exists():
        ensure_db(db_path)        # 创建新空白 KB（已含 稳定ID 列）
        return False

    with _thread_lock:
        with _FileLock(db_path):
            wb = openpyxl.load_workbook(db_path)
            ws = wb.active

            headers = [str(ws.cell(1, c).value or '').strip()
                       for c in range(1, ws.max_column + 1)]

            changed = False

            # 1. 必要时追加 稳定ID 列
            if '稳定ID' not in headers:
                col = ws.max_column + 1
                cell = ws.cell(1, col, '稳定ID')
                cell.font      = Font(bold=True, color='FFFFFF')
                cell.fill      = PatternFill(fill_type='solid', fgColor='2E74B5')
                cell.alignment = Alignment(horizontal='center')
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(col)
                ].width = 14
                stable_col = col
                headers.append('稳定ID')
                changed = True
            else:
                stable_col = headers.index('稳定ID') + 1

            # 2. 给空 稳定ID 的行回填
            for excel_row in range(2, ws.max_row + 1):
                cell = ws.cell(excel_row, stable_col)
                if cell.value is not None and str(cell.value).strip():
                    continue
                # 跳过整行空（删除痕迹）
                if all(ws.cell(excel_row, c).value is None
                       for c in range(1, ws.max_column + 1)):
                    continue
                entry = {}
                for c, h in enumerate(headers, start=1):
                    val = ws.cell(excel_row, c).value
                    entry[h] = val if val is not None else ''
                cell.value = _compute_stable_id(entry)
                changed = True

            if changed:
                _save_atomic(wb, db_path)
            return changed


if __name__ == '__main__':
    # 命令行入口：python -m core.kb_migrate /path/to/error_db.xlsx
    import sys
    if len(sys.argv) < 2:
        print('用法: python -m core.kb_migrate <error_db.xlsx 路径>')
        sys.exit(1)
    target = sys.argv[1]
    print(f'迁移目标：{target}')
    result = run(target)
    print(f'迁移完成，文件{"已修改" if result else "无需变更"}')
