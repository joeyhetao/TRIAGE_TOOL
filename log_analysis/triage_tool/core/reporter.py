# -*- coding: utf-8 -*-
from pathlib import Path
from datetime import datetime
from html import escape as h
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── 颜色常量 ────────────────────────────────────────────
COLORS = {
    'UVM_FATAL':   'C00000',
    'UVM_ERROR':   'ED7D31',
    'UVM_WARNING': 'FFD966',
    'matched':     '70AD47',
    'unmatched':   'A6A6A6',
    'header':      '2E74B5',
    'subheader':   'D6E4F0',
}

LEVEL_HTML = {
    'UVM_FATAL':   '#C00000',
    'UVM_ERROR':   '#ED7D31',
    'UVM_WARNING': '#FFD966',
}


def _header_style():
    return (
        Font(bold=True, color='FFFFFF'),
        PatternFill(fill_type='solid', fgColor=COLORS['header']),
        Alignment(horizontal='center', vertical='center', wrap_text=True),
    )


def _thin_border():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)


def generate_excel(results: list, output_path: str) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    summary_data = []

    for r in results:
        ws = wb.create_sheet(title=Path(r['file']).stem[:28])
        _write_log_sheet(ws, r)
        top = r.get('top_errors', [])
        fe = top[0] if top else {}
        match = r.get('match', {})
        entry = match.get('entry') or {}
        summary_data.append([
            r['file'],
            r['statistics']['UVM_FATAL'],
            r['statistics']['UVM_ERROR'],
            r['statistics']['UVM_WARNING'],
            fe.get('error_id', '-'),
            fe.get('level', '-'),
            '命中' if match.get('status') == 'matched' else '未匹配',
            entry.get('报错原因', '-'),
        ])

    ws_sum = wb.create_sheet(title='汇总', index=0)
    _write_summary_sheet(ws_sum, summary_data)

    wb.save(output_path)
    return output_path


def _write_log_sheet(ws, r: dict):
    border = _thin_border()

    def w(row_data, fill_color=None, bold=False):
        ws.append(row_data)
        row_idx = ws.max_row
        for cell in ws[row_idx]:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            if fill_color:
                cell.fill = PatternFill(fill_type='solid', fgColor=fill_color)
            if bold:
                cell.font = Font(bold=True)

    # 标题
    ws.append([f'日志文件：{r["file"]}'])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    title_cell = ws['A1']
    title_cell.font = Font(bold=True, size=13, color='1F497D')
    title_cell.alignment = Alignment(horizontal='center')
    ws.append([])

    # 错误统计
    w(['错误统计', '', '', ''], fill_color='D6E4F0', bold=True)
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=4)
    stat = r['statistics']
    w(['UVM_FATAL', stat['UVM_FATAL'], 'UVM_ERROR', stat['UVM_ERROR']])
    w(['UVM_WARNING', stat['UVM_WARNING'], '合计',
       stat['UVM_FATAL'] + stat['UVM_ERROR'] + stat['UVM_WARNING']])
    ws.append([])

    # 前 N 条错误及匹配结果
    top_errors = r.get('top_errors', [])
    if top_errors:
        for idx, err in enumerate(top_errors, 1):
            w([f'错误 #{idx}', '', '', ''], fill_color='D6E4F0', bold=True)
            ws.merge_cells(start_row=ws.max_row, start_column=1,
                           end_row=ws.max_row, end_column=4)
            level_color = COLORS.get(err['level'], 'FFFFFF')
            w(['错误级别', err['level'], '时间戳', err['timestamp']],
              fill_color=level_color)
            w(['错误ID', err['error_id'], '位置', err['location']])
            w(['描述', err['description'], '', ''])

            ematch = err.get('match', {})
            if ematch.get('status') == 'matched':
                eentry = ematch['entry']
                w(['匹配状态', '✅ 命中知识库', '匹配方式',
                   'ID精确匹配' if ematch.get('match_by') == 'error_id' else '关键词匹配'],
                  fill_color='E2EFDA')
                w(['报错原因', eentry.get('报错原因', ''), '根因分类', eentry.get('根因分类', '')])
                w(['所属模块', eentry.get('所属模块', ''), '录入人',   eentry.get('录入人', '')])
                w(['解决方案', eentry.get('解决方案', ''), '', ''])
                w(['关联用例', eentry.get('关联用例', ''), '', ''])
            elif ematch.get('status') == 'unmatched':
                w(['匹配状态', '❌ 未匹配 — 建议将该报错条目添加至错误数据库', '', ''],
                  fill_color='F2F2F2')
            ws.append([])
    else:
        w(['无报错信息', '', '', ''])

    for col, width in zip(['A', 'B', 'C', 'D'], [16, 40, 16, 40]):
        ws.column_dimensions[col].width = width


def _write_summary_sheet(ws, summary_data: list):
    font_h, fill_h, align_h = _header_style()
    border = _thin_border()

    ws.append(['仿真日志分析汇总'])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws['A1'].font = Font(bold=True, size=14, color='1F497D')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append([])

    headers = ['日志文件', 'FATAL数', 'ERROR数', 'WARNING数',
               '首错ID', '首错级别', '匹配状态', '报错原因']
    ws.append(headers)
    for cell in ws[3]:
        cell.font = font_h
        cell.fill = fill_h
        cell.alignment = align_h
        cell.border = border

    for row in summary_data:
        ws.append(row)
        row_idx = ws.max_row
        fill_color = None
        if row[6] == '未匹配':
            fill_color = 'F2F2F2'
        elif row[1] > 0:
            fill_color = 'FFE0E0'
        for cell in ws[row_idx]:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            if fill_color:
                cell.fill = PatternFill(fill_type='solid', fgColor=fill_color)

    for col, width in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'],
                          [30, 10, 10, 12, 20, 14, 12, 40]):
        ws.column_dimensions[col].width = width


# ── HTML报告 ─────────────────────────────────────────────
def generate_html(results: list, output_path: str) -> str:
    total_fatal   = sum(r['statistics']['UVM_FATAL']   for r in results)
    total_error   = sum(r['statistics']['UVM_ERROR']   for r in results)
    total_warning = sum(r['statistics']['UVM_WARNING'] for r in results)
    total_unmatched = sum(
        1 for r in results if r.get('match', {}).get('status') == 'unmatched'
    )

    log_sections = ''
    for r in results:
        top_errors = r.get('top_errors', [])
        stat = r['statistics']

        errors_html = ''
        for idx, err in enumerate(top_errors, 1):
            level = err.get('level', '')
            level_color = LEVEL_HTML.get(level, '#CCCCCC')
            ematch = err.get('match', {})
            eentry = ematch.get('entry') or {}

            if ematch.get('status') == 'matched':
                by = ('ID精确匹配' if ematch.get('match_by') == 'error_id'
                      else ('关键词匹配' if ematch.get('match_by') == 'keywords'
                            else '手动录入'))
                match_html = f'''
              <div class="match-box matched">
                <div class="match-title">✅ 命中知识库 <span class="match-by">（{h(by)}）</span></div>
                <table class="info-table">
                  <tr><td>报错原因</td><td>{h(str(eentry.get("报错原因","")))}</td>
                      <td>根因分类</td><td>{h(str(eentry.get("根因分类","")))}</td></tr>
                  <tr><td>所属模块</td><td>{h(str(eentry.get("所属模块","")))}</td>
                      <td>录入人</td><td>{h(str(eentry.get("录入人","")))}</td></tr>
                  <tr><td>解决方案</td><td colspan="3">{h(str(eentry.get("解决方案","")))}</td></tr>
                  <tr><td>关联用例</td><td colspan="3">{h(str(eentry.get("关联用例","")))}</td></tr>
                </table>
              </div>'''
            elif ematch.get('status') == 'unmatched':
                match_html = '''
              <div class="match-box unmatched">
                ❌ 未匹配 — 建议将该报错条目添加至错误数据库
              </div>'''
            else:
                match_html = '<div class="match-box">— 无匹配信息</div>'

            errors_html += f'''
          <div class="error-item">
            <div class="error-item-hdr" style="border-left:4px solid {level_color}">
              <span class="error-idx">#{idx}</span>
              <span class="badge" style="background:{level_color}">{h(level)}</span>
              <span class="error-id-sm">[{h(err.get("error_id",""))}]</span>
              <span class="error-ts">@ {h(err.get("timestamp",""))}</span>
            </div>
            <table class="info-table">
              <tr><td>位置</td><td colspan="3" class="desc">{h(err.get("location",""))}</td></tr>
              <tr><td>描述</td><td colspan="3" class="desc">{h(err.get("description",""))}</td></tr>
            </table>
            {match_html}
          </div>'''

        if not top_errors:
            errors_html = '<div class="match-box">— 无报错信息</div>'

        log_sections += f'''
        <div class="log-card">
          <div class="log-header">
            <span class="log-name">📄 {h(r["file"])}</span>
            <span class="stat-badges">
              <span class="sbadge fatal">FATAL: {stat["UVM_FATAL"]}</span>
              <span class="sbadge error">ERROR: {stat["UVM_ERROR"]}</span>
              <span class="sbadge warning">WARNING: {stat["UVM_WARNING"]}</span>
            </span>
          </div>
          <div class="log-body">
            <div class="section-title">前 {len(top_errors)} 条错误</div>
            {errors_html}
          </div>
        </div>'''

    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8">
  <title>仿真日志分析报告</title>
  <style>
    body{{font-family:"Microsoft YaHei",Arial,sans-serif;background:#F4F6F9;
         margin:0;padding:20px;color:#333}}
    h1{{color:#1F497D;text-align:center;margin-bottom:4px}}
    .subtitle{{text-align:center;color:#888;margin-bottom:20px;font-size:13px}}
    .summary-bar{{display:flex;gap:16px;justify-content:center;margin-bottom:24px}}
    .scard{{background:#fff;border-radius:8px;padding:12px 24px;text-align:center;
            box-shadow:0 1px 4px rgba(0,0,0,.1)}}
    .scard .num{{font-size:28px;font-weight:bold}}
    .scard .lbl{{font-size:12px;color:#888}}
    .fatal-num{{color:#C00000}} .error-num{{color:#ED7D31}}
    .warning-num{{color:#B8860B}} .unmatch-num{{color:#A6A6A6}}
    .log-card{{background:#fff;border-radius:8px;margin-bottom:16px;
               box-shadow:0 1px 4px rgba(0,0,0,.1);overflow:hidden}}
    .log-header{{background:#2E74B5;color:#fff;padding:10px 16px;
                 display:flex;justify-content:space-between;align-items:center}}
    .log-name{{font-weight:bold;font-size:14px}}
    .stat-badges{{display:flex;gap:8px}}
    .sbadge{{border-radius:4px;padding:2px 8px;font-size:12px;font-weight:bold}}
    .sbadge.fatal{{background:#C00000}} .sbadge.error{{background:#ED7D31}}
    .sbadge.warning{{background:#B8860B}}
    .log-body{{padding:16px}}
    .section-title{{font-weight:bold;color:#2E74B5;margin:12px 0 6px;
                    font-size:13px;border-bottom:1px solid #E0E0E0;padding-bottom:4px}}
    .error-item{{background:#FAFAFA;border:1px solid #EEE;border-radius:4px;
                 padding:10px;margin-bottom:10px}}
    .error-item-hdr{{display:flex;align-items:center;gap:8px;padding:4px 8px;
                     margin:-10px -10px 8px;background:#F5F5F5;border-radius:4px 4px 0 0}}
    .error-idx{{font-weight:bold;color:#555;font-size:13px}}
    .error-id-sm{{font-family:monospace;font-size:12px;color:#666}}
    .error-ts{{font-size:12px;color:#999}}
    .info-table{{width:100%;border-collapse:collapse;font-size:13px;margin-bottom:8px}}
    .info-table td{{padding:5px 8px;border:1px solid #E8E8E8}}
    .info-table td:nth-child(odd){{background:#F5F8FF;font-weight:bold;
                                   width:12%;white-space:nowrap}}
    .desc{{word-break:break-all}}
    .badge{{border-radius:4px;padding:2px 8px;color:#fff;font-size:12px;font-weight:bold}}
    .match-box{{border-radius:4px;padding:10px 14px;font-size:13px;margin-top:4px}}
    .match-box.matched{{background:#E8F5E9;border-left:4px solid #70AD47}}
    .match-box.unmatched{{background:#F5F5F5;border-left:4px solid #A6A6A6;color:#666}}
    .match-title{{font-weight:bold;color:#2E7D32;margin-bottom:8px}}
    .match-by{{font-weight:normal;color:#888;font-size:12px}}
  </style>
</head>
<body>
  <h1>仿真日志分析报告</h1>
  <div class="subtitle">生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} &nbsp;|&nbsp; 共分析 {len(results)} 个日志文件</div>
  <div class="summary-bar">
    <div class="scard"><div class="num">{len(results)}</div><div class="lbl">日志总数</div></div>
    <div class="scard"><div class="num fatal-num">{total_fatal}</div><div class="lbl">UVM_FATAL</div></div>
    <div class="scard"><div class="num error-num">{total_error}</div><div class="lbl">UVM_ERROR</div></div>
    <div class="scard"><div class="num warning-num">{total_warning}</div><div class="lbl">UVM_WARNING</div></div>
    <div class="scard"><div class="num unmatch-num">{total_unmatched}</div><div class="lbl">未匹配</div></div>
  </div>
  {log_sections}
</body>
</html>'''

    Path(output_path).write_text(html, encoding='utf-8')
    return output_path
