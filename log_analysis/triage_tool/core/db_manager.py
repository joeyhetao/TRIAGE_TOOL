# -*- coding: utf-8 -*-
from pathlib import Path
from datetime import date
import os
import time
import threading
import openpyxl

# ── 锁机制说明 ────────────────────────────────────────────
# _thread_lock : 进程内线程锁，序列化同一进程内的并发写
# _FileLock    : 跨进程/跨机器文件锁（基于 .lock 文件 + O_EXCL 原子创建）
#                适配知识库存放在网络共享盘、多台机器同时写入的场景
#                仅用标准库实现，无第三方依赖

_thread_lock = threading.Lock()


class _FileLock:
    """
    基于 .lock 文件的跨进程文件锁。
    - O_CREAT|O_EXCL 保证创建锁文件的原子性（NTFS 及主流网络文件系统均支持）
    - 自动清理超时僵尸锁（进程崩溃后遗留的锁文件）
    """
    STALE_TIMEOUT = 60   # 锁文件超过此秒数视为僵尸锁，自动清除

    def __init__(self, db_path: str, timeout: float = 15, retry: float = 0.05):
        self.lock_path = str(db_path) + '.lock'
        self.timeout   = timeout
        self.retry     = retry
        self._fd       = None

    def __enter__(self):
        deadline = time.monotonic() + self.timeout
        while True:
            try:
                self._fd = os.open(
                    self.lock_path,
                    os.O_CREAT | os.O_EXCL | os.O_WRONLY | getattr(os, 'O_BINARY', 0)
                )
                return self
            except FileExistsError:
                # 检查是否为僵尸锁（持锁进程崩溃后未释放）
                try:
                    age = time.time() - os.path.getmtime(self.lock_path)
                    if age > self.STALE_TIMEOUT:
                        try:
                            os.remove(self.lock_path)
                        except OSError:
                            pass  # Windows: 文件仍被持有，等待自然释放
                        continue
                except FileNotFoundError:
                    continue  # 锁文件刚被别人释放，重试
                if time.monotonic() > deadline:
                    raise TimeoutError(
                        f'等待知识库锁超时（>{self.timeout}s），'
                        f'请检查 {self.lock_path} 是否残留'
                    )
                time.sleep(self.retry)

    def __exit__(self, *_):
        if self._fd is not None:
            os.close(self._fd)
            self._fd = None
        try:
            os.remove(self.lock_path)
        except FileNotFoundError:
            pass


HEADERS = [
    '错误类型', '错误ID', '关键描述关键词', '报错原因',
    '所属模块', '根因分类', '解决方案', '关联用例',
    '录入人', '录入日期',
]


def ensure_db(db_path: str) -> None:
    """若知识库不存在则自动创建含表头的空白Excel文件。"""
    path = Path(db_path)
    if path.exists():
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '错误知识库'
    ws.append(HEADERS)
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='2E74B5')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    col_widths = [14, 18, 30, 30, 14, 14, 30, 25, 10, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    wb.save(str(path))


def load_db(db_path: str) -> list:
    """读取知识库，返回字典列表。每条记录含 _row_idx（Excel 行号，从2开始）。
    读取失败时最多重试3次（防止写入期间的短暂竞争）。"""
    ensure_db(db_path)
    last_err = None
    for attempt in range(3):
        try:
            wb = openpyxl.load_workbook(db_path)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                return []
            headers = [str(h).strip() if h else '' for h in rows[0]]
            entries = []
            for excel_row, row in enumerate(rows[1:], start=2):
                if all(cell is None for cell in row):
                    continue
                entry = {headers[i]: (row[i] if row[i] is not None else '')
                         for i in range(len(headers))}
                entry['_row_idx'] = excel_row
                entries.append(entry)
            return entries
        except Exception as e:
            last_err = e
            time.sleep(0.1 * (attempt + 1))
    raise RuntimeError(f'读取知识库失败（重试3次）: {last_err}')


def find_duplicates(db_path: str, entry: dict, exclude_row_idx: int = None) -> list:
    """
    检查 entry 是否与知识库中已有条目重复（不检查录入人）。
    判断规则（满足其一即视为重复）：
      - 错误类型相同 AND 错误ID相同（两者均非空，忽略大小写）
      - 错误类型相同 AND 关键描述关键词相同（两者均非空，标准化逗号/空格后比较）
      - 错误类型相同 AND 报错原因相同（两者均非空，忽略首尾空白）
      - 错误类型相同 AND 解决方案相同（两者均非空，忽略首尾空白）
    exclude_row_idx: 编辑时排除自身行，避免与自己重复。
    返回所有冲突条目列表（含 _row_idx）。
    """
    import re as _re
    def _norm_kw(s):
        return _re.sub(r'\s*[,，]\s*', ',', str(s or '').strip().lower()).strip(',')

    level    = str(entry.get('错误类型', '') or '').strip().upper()
    error_id = str(entry.get('错误ID',   '') or '').strip().lower()
    kw_norm  = _norm_kw(entry.get('关键描述关键词', ''))
    reason   = str(entry.get('报错原因',  '') or '').strip()
    solution = str(entry.get('解决方案',  '') or '').strip()

    conflicts = []
    for e in load_db(db_path):
        if exclude_row_idx and e.get('_row_idx') == exclude_row_idx:
            continue
        if str(e.get('错误类型', '') or '').strip().upper() != level:
            continue
        e_id       = str(e.get('错误ID',   '') or '').strip().lower()
        e_kw       = _norm_kw(e.get('关键描述关键词', ''))
        e_reason   = str(e.get('报错原因',  '') or '').strip()
        e_solution = str(e.get('解决方案',  '') or '').strip()
        id_dup       = bool(error_id and e_id       and error_id == e_id)
        kw_dup       = bool(kw_norm  and e_kw       and kw_norm  == e_kw)
        reason_dup   = bool(reason   and e_reason   and reason   == e_reason)
        solution_dup = bool(solution and e_solution and solution  == e_solution)
        if id_dup or kw_dup or reason_dup or solution_dup:
            conflicts.append(e)
    return conflicts


def update_entry(db_path: str, row_idx: int, new_data: dict) -> None:
    """
    更新知识库中指定行（row_idx 为 Excel 行号，从2开始）。
    只更新 HEADERS 中定义的字段，忽略 _row_idx 等内部字段。
    """
    with _thread_lock:
        with _FileLock(db_path):
            wb = openpyxl.load_workbook(db_path)
            ws = wb.active
            headers = [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value else ''
                       for c in range(1, len(HEADERS) + 1)]
            for col_idx, header in enumerate(headers, 1):
                if header in new_data:
                    ws.cell(row=row_idx, column=col_idx, value=new_data[header])
            wb.save(db_path)


def delete_entry(db_path: str, row_idx: int) -> None:
    """删除知识库中指定行（row_idx 为 Excel 行号，从2开始）。"""
    with _thread_lock:
        with _FileLock(db_path):
            wb = openpyxl.load_workbook(db_path)
            ws = wb.active
            ws.delete_rows(row_idx)
            wb.save(db_path)


def append_entry(db_path: str, entry: dict) -> None:
    """
    向知识库追加一条新记录，自动填写录入日期。
    线程安全 + 跨进程/跨网络共享盘安全。
    """
    entry['录入日期'] = str(date.today())
    row = [entry.get(h, '') for h in HEADERS]
    with _thread_lock:              # 进程内线程串行
        with _FileLock(db_path):    # 跨进程/跨机器串行
            ensure_db(db_path)
            wb = openpyxl.load_workbook(db_path)
            ws = wb.active
            ws.append(row)
            wb.save(db_path)
